#!/usr/bin/env python3
"""
esg-exclusions-enricher.py — Exclusions sectorielles ESG depuis les fichiers EET
=================================================================================
Alimente investissement_funds.esg_exclusions (jsonb {clé: bool}) depuis les
European ESG Templates (EET, FinDatEx) publiés par les sociétés de gestion :
champs d'exclusions PAB/CTB (tabac, armes controversées, charbon/fossiles) et
politiques d'exclusion volontaires (jeux, alcool, nucléaire…).

Pourquoi un parseur de FICHIERS plutôt qu'un crawler : l'EET n'a pas de point de
diffusion central public — chaque SGP publie son fichier (page « informations
durabilité » SFDR, doc center, ou plateformes type fundinfo). On récupère les
fichiers (CSV/XLSX) à la main ou par script dédié, puis ce script les ingère de
façon uniforme. Les en-têtes EET varient selon la version du template
(V1.0/V1.1.x : `20010_Financial_Instrument_Identifying_Data`, etc.) → détection
par MOTIFS sur les noms de colonnes plutôt que par liste figée.

Clés canoniques écrites (cf. COMMENT SQL de la colonne) :
    tobacco, controversial_weapons, weapons, thermal_coal, fossil, nuclear,
    gambling, alcohol, adult_entertainment, ungc_violations

Sémantique : clé présente = politique documentée (true = le fonds exclut le
secteur, false = il ne l'exclut pas) ; clé absente = inconnu. Le moteur
d'allocation replie sur le proxy « labels » quand la clé est absente.

Fill-only : n'écrit que les fonds DÉJÀ en base (update_funds_bulk, jamais
d'insert) et, par défaut, seulement ceux SANS esg_exclusions (--overwrite pour
rafraîchir, ex. nouvelle période EET).

Le MÊME run remplit aussi les 3 champs de durabilité MiFID depuis les colonnes
EET dédiées (cf. docs/mapping-eet-mifid.md et SUSTAIN_PATTERNS) :
sustainable_investment_pct, taxonomy_alignment_pct, pai_considered — strictement
fill-only (jamais d'--overwrite : préserve les valeurs de l'annexe Morningstar),
avec sustainability_source = 'eet:<sgp>:<période>' posé sur les fonds enrichis.

Usage :
    python3 scripts/scrapers/esg-exclusions-enricher.py --eet amundi-eet.csv            # dry-run
    python3 scripts/scrapers/esg-exclusions-enricher.py --dir data/eet/ --apply
    python3 scripts/scrapers/esg-exclusions-enricher.py --eet f.xlsx --source amundi \\
        --as-of 2026-06-30 --apply --overwrite
"""

import re
import csv
import sys
import json
import argparse
from datetime import datetime, timezone, date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from db import get_client, update_funds_bulk, log_run, now_iso

# ─── Détection des colonnes EET ───────────────────────────────────────────────

ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")

# Une colonne est candidate « exclusion » si son en-tête évoque une politique
# d'exclusion/de filtrage (les champs EET PAB/CTB et les politiques volontaires
# contiennent l'un de ces marqueurs), pour ne pas confondre avec les colonnes
# d'EXPOSITION (% d'implication) ou de simple description.
POLICY_MARKER = re.compile(r"exclu|policy|polic|screen|\bpab\b|\bctb\b|_pab_|_ctb_", re.IGNORECASE)

# Motif d'en-tête → clé canonique. L'ordre compte (controversial weapons avant
# weapons génériques).
SECTOR_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"controversial[\s_]*weapon", re.I), "controversial_weapons"),
    (re.compile(r"weapon|armament|defen[cs]e", re.I), "weapons"),
    (re.compile(r"tobacco", re.I),                    "tobacco"),
    (re.compile(r"thermal[\s_]*coal|coal", re.I),     "thermal_coal"),
    (re.compile(r"fossil", re.I),                     "fossil"),
    (re.compile(r"nuclear", re.I),                    "nuclear"),
    (re.compile(r"gambling", re.I),                   "gambling"),
    (re.compile(r"alcohol", re.I),                    "alcohol"),
    (re.compile(r"adult|pornograph", re.I),           "adult_entertainment"),
    (re.compile(r"global[\s_]*compact|ungc|norms?[\s_]*based", re.I), "ungc_violations"),
]

TRUE_VALUES  = {"y", "yes", "true", "1", "oui"}
FALSE_VALUES = {"n", "no", "false", "0", "non"}

# ─── Durabilité MiFID (mêmes fichiers EET, cf. docs/mapping-eet-mifid.md) ─────
# Trois colonnes quantitatives de investissement_funds remplies depuis l'EET.
# Motifs par PRIORITÉ décroissante : code FinDatEx V1.1.x d'abord (observé dans
# les fichiers réels — les codes « usuels » 20510/20530/20440 du doc désignaient
# d'autres champs), libellé générique en repli (variantes de version).
#   sustainable_investment_pct : 20420 = part minimale/planifiée TOTALE d'inv.
#     durables (art. 2(17)) ; replis 20180 (bloc art. 8) puis 20220 (art. 9,
#     objectif environnemental — plancher conservateur si le social s'y ajoute).
#   taxonomy_alignment_pct : 20610 = % minimal aligné taxinomie HORS dette
#     souveraine (la variante affichée dans l'annexe) ; replis 20600 (incl.
#     souverain) puis 20450 (bloc SFDR).
#   pai_considered : 20100 (Y/N).
SUSTAIN_PATTERNS: dict[str, list[re.Pattern]] = {
    "sustainable_investment_pct": [
        re.compile(r"^20420_|minimum_or_planned_investments_sustainable_investments$", re.I),
        re.compile(r"^20180_|minimal_proportion_of_sustainable_investments", re.I),
        re.compile(r"^20220_|minimum_sustainable_investment_with_environmental", re.I),
    ],
    "taxonomy_alignment_pct": [
        re.compile(r"^20610_|aligned_eu_taxonomy_excl_sovereign_bonds$", re.I),
        re.compile(r"^20600_|aligned_eu_taxonomy_incl_sovereign_bonds$", re.I),
        re.compile(r"^20450_|sustainable_investments_taxonomy_aligned$", re.I),
    ],
    "pai_considered": [
        re.compile(r"^20100_|consider.*(?:principle|principal).*adverse.*impact", re.I),
    ],
}


def parse_pct(raw) -> float | None:
    """Valeur EET → pourcentage 0-100, ou None. L'EET exprime les parts en
    FRACTION (0.20 = 20 %) ; certains producteurs écrivent déjà en % → règle du
    doc : v ≤ 1 → ×100. Bornes 0-100 strictes, arrondi 2 décimales."""
    if raw is None:
        return None
    s = str(raw).strip().replace(",", ".").replace("%", "")
    if not s:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    if 0 <= v <= 1:
        v *= 100
    if not (0 <= v <= 100):
        return None
    return round(v, 2)


def map_sustain_headers(headers: list[str]) -> dict[str, list[int]]:
    """{colonne DB: [index d'en-tête par priorité décroissante]}."""
    out: dict[str, list[int]] = {}
    for field, patterns in SUSTAIN_PATTERNS.items():
        for pat in patterns:
            for i, h in enumerate(headers):
                if h and pat.search(str(h).strip()):
                    out.setdefault(field, []).append(i)
                    break  # une colonne par niveau de priorité
    return out


def parse_sustain_row(row: list, cols: dict[str, list[int]]) -> dict:
    """Champs durabilité d'une ligne EET : première priorité renseignée.
    Garde-fou plausibilité : la part taxinomie est un sous-ensemble de l'inv.
    durable → si taxo > SI (+ epsilon), on jette la taxo (valeur douteuse)."""
    out: dict = {}
    for field, indexes in cols.items():
        for i in indexes:
            raw = row[i] if i < len(row) else None
            v = parse_bool(raw) if field == "pai_considered" else parse_pct(raw)
            if v is not None:
                out[field] = v
                break
    si, taxo = out.get("sustainable_investment_pct"), out.get("taxonomy_alignment_pct")
    if si is not None and taxo is not None and taxo > si + 0.01:
        out.pop("taxonomy_alignment_pct", None)
    return out


def parse_bool(raw) -> bool | None:
    """Y/N/True/False/1/0/Oui/Non → bool ; tout le reste (vide, %, texte) → None."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in TRUE_VALUES:
        return True
    if s in FALSE_VALUES:
        return False
    return None


def map_headers(headers: list[str]) -> tuple[int | None, dict[int, str]]:
    """
    Repère la colonne ISIN et les colonnes d'exclusion.
    Retourne (index_isin, {index_colonne: clé canonique}).
    """
    isin_idx: int | None = None
    sector_cols: dict[int, str] = {}
    for i, h in enumerate(headers):
        name = (h or "").strip()
        if not name:
            continue
        if isin_idx is None and re.search(r"\bisin\b|identifying[\s_]*data|instrument[\s_]*identif", name, re.I):
            isin_idx = i
            continue
        if not POLICY_MARKER.search(name):
            continue
        for pat, key in SECTOR_PATTERNS:
            if pat.search(name):
                sector_cols[i] = key
                break
    return isin_idx, sector_cols


# ─── Lecture des fichiers (CSV / XLSX) ────────────────────────────────────────

def read_rows(path: Path) -> list[list]:
    """Fichier EET → liste de lignes (la première contenant les en-têtes)."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit("openpyxl requis pour les .xlsx — pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]

    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = raw[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        # EET : le point-virgule est le séparateur le plus répandu.
        class dialect:  # type: ignore[no-redef]
            delimiter = ";"
            quotechar = '"'
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL
            doublequote = True
            skipinitialspace = False
    return [row for row in csv.reader(raw.splitlines(), dialect=dialect)]


def find_isin_by_values(rows: list[list]) -> int | None:
    """Repli : colonne dont les valeurs RESSEMBLENT à des ISIN (en-tête atypique)."""
    if len(rows) < 2:
        return None
    width = max(len(r) for r in rows[1:21])
    for i in range(width):
        vals = [str(r[i]).strip().upper() for r in rows[1:21] if i < len(r) and r[i]]
        if vals and sum(1 for v in vals if ISIN_RE.match(v)) >= max(2, len(vals) // 2):
            return i
    return None


def parse_eet_file(
    path: Path,
) -> tuple[dict[str, dict[str, bool]], dict[str, str], dict[str, dict]]:
    """
    Parse un fichier EET → ({isin: {clé: bool}}, {clé: en-tête source},
    {isin: champs durabilité MiFID}).
    Plusieurs colonnes pour une même clé d'exclusion : OR (une politique
    documentée suffit) ; seules les valeurs booléennes explicites sont retenues
    (les seuils en % et les champs descriptifs sont ignorés). Les champs
    durabilité (SI %, taxo %, PAI) sont lus par priorité de colonne
    (cf. SUSTAIN_PATTERNS).
    """
    rows = read_rows(path)
    if not rows:
        return {}, {}, {}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    isin_idx, sector_cols = map_headers(headers)
    if isin_idx is None:
        isin_idx = find_isin_by_values(rows)
    sustain_cols = map_sustain_headers(headers)
    if isin_idx is None or (not sector_cols and not sustain_cols):
        return {}, {}, {}

    mapping_info = {key: headers[i] for i, key in sector_cols.items()}
    out: dict[str, dict[str, bool]] = {}
    sustain: dict[str, dict] = {}
    for row in rows[1:]:
        if isin_idx >= len(row):
            continue
        isin = str(row[isin_idx] or "").strip().upper()
        if not ISIN_RE.match(isin):
            continue
        excl = out.setdefault(isin, {})
        for i, key in sector_cols.items():
            val = parse_bool(row[i]) if i < len(row) else None
            if val is None:
                continue
            # OR entre colonnes d'une même clé (PAB + politique volontaire, etc.)
            excl[key] = excl.get(key, False) or val
        if not excl:
            out.pop(isin, None)
        s = parse_sustain_row(row, sustain_cols)
        if s:
            # Plusieurs parts d'un même ISIN : on complète sans écraser.
            cur = sustain.setdefault(isin, {})
            for k, v in s.items():
                cur.setdefault(k, v)
    return out, mapping_info, sustain


# ─── Main ─────────────────────────────────────────────────────────────────────

def run(files: list[Path], apply: bool, overwrite: bool, limit: int | None,
        source: str | None, as_of: str | None):
    print("=" * 64)
    print("  ESG Exclusions Enricher — ingestion EET (FinDatEx)")
    print("=" * 64)
    print(f"  Mode      : {'APPLY' if apply else 'DRY-RUN'}{' + OVERWRITE' if overwrite else ''}")
    print(f"  Fichiers  : {len(files)}")
    print()

    started = datetime.now(timezone.utc)
    as_of_date = as_of or date.today().isoformat()

    merged: dict[str, dict[str, bool]] = {}
    sources: dict[str, str] = {}
    sustain_all: dict[str, dict] = {}
    sustain_src: dict[str, str] = {}
    for path in files:
        data, mapping, sustain = parse_eet_file(path)
        label = f"eet:{source or path.stem}:{as_of_date}"
        print(f"  {path.name} : {len(data)} ISIN excl., {len(sustain)} ISIN durabilité")
        if not data and not sustain:
            print("    ⚠️  aucune colonne ISIN/exclusion/durabilité reconnue — en-têtes inattendus ?")
            continue
        for key, header in sorted(mapping.items()):
            print(f"    {key:<22} ← {header}")
        for isin, excl in data.items():
            merged.setdefault(isin, {}).update(excl)
            sources[isin] = label
        for isin, s in sustain.items():
            cur = sustain_all.setdefault(isin, {})
            for k, v in s.items():
                cur.setdefault(k, v)
            sustain_src.setdefault(isin, label)

    if not merged and not sustain_all:
        print("\n  ⚠️  Aucune donnée extraite")
        return
    if limit:
        merged = dict(list(merged.items())[:limit])
        sustain_all = {k: v for k, v in sustain_all.items() if k in merged}

    # Statistiques par clé
    counts: dict[str, int] = {}
    for excl in merged.values():
        for k, v in excl.items():
            if v:
                counts[k] = counts.get(k, 0) + 1
    print(f"\n  {len(merged)} ISIN au total — exclusions positives par clé :")
    for k, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"    {k:<22} {n}")
    s_counts = {f: sum(1 for s in sustain_all.values() if f in s) for f in SUSTAIN_PATTERNS}
    print(f"  {len(sustain_all)} ISIN avec durabilité MiFID : "
          + "  ".join(f"{f.replace('_pct','')}:{n}" for f, n in s_counts.items()))

    # Filtrage : ISINs déjà en base, et (sauf --overwrite) sans donnée existante.
    # Les champs durabilité MiFID sont STRICTEMENT fill-only (jamais d'overwrite :
    # ~200 valeurs posées par l'annexe Morningstar à préserver).
    client = get_client()
    isins = sorted(set(merged) | set(sustain_all))
    existing: set[str] = set()
    already: set[str] = set()
    sustain_null: dict[str, set[str]] = {}  # isin → champs MiFID encore NULL en base
    has_column = True
    for i in range(0, len(isins), 500):
        chunk = isins[i : i + 500]
        cols = ("isin, esg_exclusions, sustainable_investment_pct, "
                "taxonomy_alignment_pct, pai_considered") if has_column else "isin"
        try:
            r = client.table("investissement_funds").select(cols).in_("isin", chunk).execute()
        except Exception as e:
            # Migration 20260721160000 pas encore appliquée : on peut dry-runner,
            # mais pas écrire.
            if has_column and "esg_exclusions" in str(e):
                has_column = False
                print("  ⚠️  Colonne esg_exclusions absente en base — appliquer la migration 20260721160000 avant --apply.")
                r = client.table("investissement_funds").select("isin").in_("isin", chunk).execute()
            else:
                raise
        for row in (r.data or []):
            existing.add(row["isin"])
            if row.get("esg_exclusions") is not None:
                already.add(row["isin"])
            nulls = {f for f in SUSTAIN_PATTERNS if row.get(f) is None}
            if nulls:
                sustain_null[row["isin"]] = nulls

    batch: list[dict] = []
    n_excl = n_sustain = 0
    sustain_written = {f: 0 for f in SUSTAIN_PATTERNS}
    for isin in isins:
        row: dict = {"isin": isin}
        if isin in merged and isin in existing and (overwrite or isin not in already):
            row.update({
                "esg_exclusions": merged[isin],
                "esg_exclusions_source": sources[isin],
                "esg_exclusions_updated_at": as_of_date,
            })
            n_excl += 1
        # Durabilité MiFID : uniquement les champs trouvés ET NULL en base.
        keep = {f: v for f, v in sustain_all.get(isin, {}).items()
                if f in sustain_null.get(isin, set())}
        if keep and isin in existing:
            row.update(keep)
            row["sustainability_source"] = sustain_src[isin]
            row["sustainability_computed_at"] = now_iso()
            n_sustain += 1
            for f in keep:
                sustain_written[f] += 1
        if len(row) > 1:
            batch.append(row)

    print(f"\n  En base : {len(existing)}/{len(isins)} — "
          f"exclusions déjà renseignées : {len(already)}"
          f"{' (écrasées)' if overwrite else ' (conservées, --overwrite pour rafraîchir)'}")
    print(f"  À écrire : {len(batch)} fonds — exclusions : {n_excl}, durabilité MiFID : {n_sustain} "
          f"({'  '.join(f'{f}:{n}' for f, n in sustain_written.items())})")

    if not apply:
        print("\n  Aperçu (10 premiers) :")
        for row in batch[:10]:
            payload = {k: v for k, v in row.items() if k != "isin"}
            print(f"  {row['isin']} → {json.dumps(payload, sort_keys=True, default=str)}")
        print("\n  DRY-RUN — relancer avec --apply pour écrire.")
        return

    if not has_column:
        print("\n  ✗ ABANDON : la colonne esg_exclusions n'existe pas encore en base.")
        return

    ok, fail = update_funds_bulk(batch, batch_size=200)
    print(f"  → Update {len(batch)} fonds : {ok} OK, {fail} échec")
    log_run("esg-exclusions-enricher", "success" if fail == 0 else "partial",
            ok, fail, started_at=started)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Exclusions ESG depuis fichiers EET")
    parser.add_argument("--eet", action="append", default=[], help="Fichier EET (CSV/XLSX), répétable")
    parser.add_argument("--dir", type=str, help="Dossier de fichiers EET (*.csv, *.xlsx)")
    parser.add_argument("--apply", action="store_true", help="Écrire dans Supabase")
    parser.add_argument("--overwrite", action="store_true", help="Écraser les esg_exclusions existants")
    parser.add_argument("--limit", type=int, help="Limiter à N fonds")
    parser.add_argument("--source", type=str, help="Étiquette SGP pour esg_exclusions_source (défaut : nom de fichier)")
    parser.add_argument("--as-of", type=str, help="Date de référence YYYY-MM-DD (défaut : aujourd'hui)")
    args = parser.parse_args()

    paths = [Path(p) for p in args.eet]
    if args.dir:
        paths += sorted(Path(args.dir).glob("*.csv")) + sorted(Path(args.dir).glob("*.xlsx"))
    paths = [p for p in paths if p.exists()]
    if not paths:
        raise SystemExit("Aucun fichier EET — passer --eet <fichier> ou --dir <dossier>")
    run(paths, apply=args.apply, overwrite=args.overwrite, limit=args.limit,
        source=args.source, as_of=args.as_of)
