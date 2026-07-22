#!/usr/bin/env python3
"""
labels-registries.py — Labels officiels (ISR / Greenfin / Finansol) depuis les registres
=========================================================================================
Les labels en base venaient de sources indirectes (368 isr, 90 finansol, 0 greenfin
au 21/07/2026) alors que les registres officiels sont publics et téléchargeables.
Ce scraper récupère les listes officielles et fusionne les labels par ISIN dans
investissement_funds.labels.

Sources (découverte du lien à chaque run — les noms de fichiers sont horodatés) :

  1. Référentiel des OPC labellisés — Banque de France (trimestriel).
     Collecte officielle auprès des organismes porteurs des labels ISR, Greenfin,
     Finansol, CIES et Relance ; un ISIN PAR PART/COMPARTIMENT (≈6 000 lignes),
     bien plus fin que les listes « un ISIN par fonds » des sites de labels.
     Page : banque-france.fr → référentiel-des-opc-labellises → lien .xlsx
     (colonnes : label, societe_de_gestion, nom_du_fonds, isin, code_AMF, lei,
     date_arrete). CIES/Relance sont ignorés (hors vocabulaire labels du screener).

  2. Liste officielle Label ISR — lelabelisr.fr (mensuelle, « seule la liste
     Excel est exhaustive »). Complète le référentiel BdF : les deux listes ne se
     recouvrent pas parfaitement (part principale vs toutes parts, délais de
     collecte). Feuille « Liste fonds », en-tête ligne 2, ISIN en colonne E.

  Greenfin (ministère) et Finansol (FAIR) ne publient pas d'export ISIN sur leurs
  sites — leurs listes officielles transitent précisément par le référentiel BdF
  (source 1), qui fait foi ici.

RÈGLES :
  - ADDITIF STRICT : labels = existants ∪ nouveaux ; on ne retire JAMAIS un label
    (même logique que merged_labels de sfdr-annex-enricher.py — un fonds sorti
    d'une liste garde son tag jusqu'à décision manuelle).
  - Fonds EXISTANTS uniquement (update_funds_bulk n'insère jamais) : les ISIN des
    registres absents de la base sont comptés puis ignorés.
  - Dry-run par défaut : n'écrit qu'avec --apply.

Usage :
    python3 scripts/scrapers/labels-registries.py            # dry-run (rapport)
    python3 scripts/scrapers/labels-registries.py --apply    # écrit dans Supabase
    python3 scripts/scrapers/labels-registries.py --apply --source bdf|isr
"""

import io
import re
import sys
import argparse
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from curl_cffi import requests as cr

sys.path.insert(0, str(Path(__file__).parent.parent))
from db import get_client, update_funds_bulk, log_run

TIMEOUT = 60

# Page du référentiel BdF (le lien .xlsx dedans change à chaque trimestre).
BDF_PAGE = ("https://www.banque-france.fr/statistiques/epargne-et-comptes-nationaux-financiers/"
            "les-autres-formes-depargne-opc-organismes-de-placement-collectifs-et-assurances/"
            "referentiel-des-opc-labellises")
BDF_HOST = "https://www.banque-france.fr"

# Pages « liste des fonds labellisés » du Label ISR (ancienne + nouvelle version —
# le site a déjà changé d'URL une fois, on tente les deux).
ISR_PAGES = (
    "https://www.lelabelisr.fr/comment-investir/fonds-labellises-1/",
    "https://www.lelabelisr.fr/comment-investir/fonds-labellises/",
)

# label du registre BdF → tag labels du screener (minuscule). CIES / RELANCE :
# hors vocabulaire du moteur (EXCLUSION_GUARANTEE_LABELS & filtres UI), ignorés.
BDF_LABEL_TAGS = {"ISR": "isr", "GREENFIN": "greenfin", "FINANSOL": "finansol"}

ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}\d$")


def _get(url: str) -> cr.Response:
    """GET avec empreinte navigateur (banque-france.fr renvoie 403 sinon)."""
    r = cr.get(url, impersonate="chrome", timeout=TIMEOUT)
    r.raise_for_status()
    return r


def _clean_isin(raw) -> str | None:
    """ISIN valide (format) ou None — écarte les codes internes type QS…/vides."""
    isin = str(raw or "").strip().upper()
    return isin if ISIN_RE.match(isin) else None


# ─── Source 1 : référentiel Banque de France ─────────────────────────────────

def fetch_bdf() -> dict[str, set[str]]:
    """{tag: {isin}} depuis le référentiel BdF des OPC labellisés."""
    html = _get(BDF_PAGE).text
    links = re.findall(r'href="([^"]*referentiel[^"]*\.xlsx[^"]*)"', html, re.IGNORECASE)
    if not links:
        raise RuntimeError("référentiel BdF : aucun lien .xlsx trouvé sur la page")
    url = links[0] if links[0].startswith("http") else BDF_HOST + links[0]
    print(f"  BdF : {url}")

    wb = openpyxl.load_workbook(io.BytesIO(_get(url).content), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().lower() for c in next(rows)]
    try:
        i_label, i_isin = header.index("label"), header.index("isin")
    except ValueError:
        raise RuntimeError(f"référentiel BdF : colonnes inattendues {header}")

    out: dict[str, set[str]] = {}
    ignored: dict[str, int] = {}
    for row in rows:
        label = str(row[i_label] or "").strip().upper()
        isin = _clean_isin(row[i_isin])
        if not isin:
            continue
        tag = BDF_LABEL_TAGS.get(label)
        if tag:
            out.setdefault(tag, set()).add(isin)
        else:
            ignored[label] = ignored.get(label, 0) + 1
    counts = {t: len(s) for t, s in sorted(out.items())}
    print(f"  BdF : {counts} — ignorés (hors vocabulaire) : {ignored}")
    return out


# ─── Source 2 : liste officielle Label ISR (lelabelisr.fr) ───────────────────

def fetch_isr() -> set[str]:
    """{isin} depuis l'export Excel mensuel de lelabelisr.fr."""
    url = None
    for page in ISR_PAGES:
        try:
            html = _get(page).text
        except Exception:
            continue
        links = re.findall(r'href="([^"]*\.xlsx[^"]*)"', html, re.IGNORECASE)
        links = [l for l in links if re.search(r"liste|fonds|label", l, re.IGNORECASE)]
        if links:
            url = links[0] if links[0].startswith("http") else "https://www.lelabelisr.fr" + links[0]
            break
    if not url:
        raise RuntimeError("lelabelisr.fr : export Excel introuvable sur les pages connues")
    print(f"  ISR : {url}")

    wb = openpyxl.load_workbook(io.BytesIO(_get(url).content), read_only=True)
    ws = wb[wb.sheetnames[0]]
    isins: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        # L'ISIN est en dernière colonne renseignée ; l'en-tête n'est pas en
        # ligne 1 (ligne d'accueil au-dessus) → on filtre au format plutôt que
        # de dépendre de la position exacte de l'en-tête.
        for cell in row:
            isin = _clean_isin(cell)
            if isin:
                isins.add(isin)
    print(f"  ISR : {len(isins)} ISIN valides")
    return isins


# ─── Fusion additive ─────────────────────────────────────────────────────────

def merged_labels(existing: list | None, new_tags: set[str]) -> list[str] | None:
    """labels fusionnés (existants ∪ nouveaux), ou None si rien à écrire.
    Additif STRICT : ne retire jamais un label existant."""
    cur = [str(x) for x in (existing or [])]
    merged = sorted(set(cur) | new_tags)
    return merged if merged != sorted(cur) else None


def run(apply: bool, source: str | None) -> None:
    print("=" * 70)
    print(f"  labels-registries — {'APPLY' if apply else 'DRY-RUN (aucune écriture)'}")
    print("=" * 70)
    started = datetime.now(timezone.utc)

    by_tag: dict[str, set[str]] = {}
    if source in (None, "bdf"):
        for tag, isins in fetch_bdf().items():
            by_tag.setdefault(tag, set()).update(isins)
    if source in (None, "isr"):
        by_tag.setdefault("isr", set()).update(fetch_isr())

    # ISIN → tags à garantir
    want: dict[str, set[str]] = {}
    for tag, isins in by_tag.items():
        for isin in isins:
            want.setdefault(isin, set()).add(tag)
    print(f"\n  {len(want)} ISIN au total dans les registres "
          f"({', '.join(f'{t}:{len(s)}' for t, s in sorted(by_tag.items()))})")

    # Fonds présents en base (update only — jamais d'insert)
    client = get_client()
    targets = sorted(want)
    funds: dict[str, list] = {}
    for i in range(0, len(targets), 400):
        chunk = client.table("investissement_funds").select("isin, labels") \
            .in_("isin", targets[i:i + 400]).execute().data or []
        funds.update({f["isin"]: f.get("labels") for f in chunk})
    print(f"  {len(funds)} présents en base, {len(want) - len(funds)} inconnus (ignorés)")

    updates: list[dict] = []
    added_per_tag = {t: 0 for t in by_tag}
    for isin, tags in want.items():
        if isin not in funds:
            continue
        merged = merged_labels(funds[isin], tags)
        if merged is None:
            continue
        for t in tags:
            if t not in [str(x) for x in (funds[isin] or [])]:
                added_per_tag[t] += 1
        updates.append({"isin": isin, "labels": merged})

    print(f"  {len(updates)} fonds à mettre à jour — nouveaux tags : "
          f"{ {t: n for t, n in sorted(added_per_tag.items())} }")

    if not apply:
        for u in updates[:10]:
            print(f"    ex: {u['isin']} → {u['labels']}")
        print("\n  Dry-run terminé. Relancer avec --apply pour écrire.")
        return

    if updates:
        print(f"\n  Écriture ({len(updates)} fonds)…", end=" ", flush=True)
        ok, fail = update_funds_bulk(updates, batch_size=200)
        print(f"✓ {ok} OK, {fail} échec")
        log_run(scraper="labels-registries", status="success",
                records_processed=ok, records_failed=fail, started_at=started)
    else:
        print("  Rien à écrire (labels déjà à jour).")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Labels ISR/Greenfin/Finansol depuis les registres officiels")
    p.add_argument("--apply", action="store_true", help="Écrire (fusion additive) dans Supabase")
    p.add_argument("--source", choices=["bdf", "isr"], help="Limiter à une source")
    args = p.parse_args()
    run(apply=args.apply, source=args.source)
